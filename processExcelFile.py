import openpyxl, os

os.chdir('/Users/guest')

workbook = openpyxl.load_workbook('example.xlsx')
print(workbook.get_sheet_names())

sheet = workbook.get_sheet_by_name('Sheet1')
print(type(sheet))
cell = sheet['A1']
print("\nThe date in A1 cell is " + str(cell.value))

print("\nThe fruit name in B1 cell is " + str(sheet['B1'].value))

#dynamically print cells

# prints cell that is in 1st row and 2nd column
print(sheet.cell(row=1, column=2))
print("\n")
for i in range(1, 8):
    print(i, sheet.cell(row=i, column=2).value)
    
