import openpyxl, os

wb = openpyxl.Workbook()
print(wb.get_sheet_names())

sheet = wb.get_sheet_by_name('Sheet')
print(sheet['A1'].value)

sheet['A1'] = 42
sheet['A2'] = 44

os.chdir('/Users/myfiles')
wb.save('data1.xlsx')

# create new sheet in workbook
sheet2 = wb.create_sheet()

print(wb.get_sheet_names())

sheet2.title = 'My New Sheet Name'
print(wb.get_sheet_names())
wb.save('data2.xlsx')

wb.create_sheet(index=0, title='My Other Sheet')
wb.save('data3.xlsx')

